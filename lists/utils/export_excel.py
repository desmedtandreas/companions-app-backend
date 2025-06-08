import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from companies.models import Company


def generate_companies_excel(companies, listname):
    """Generate an Excel file with company information."""
    if hasattr(companies, "prefetch_related"):
        companies = companies.prefetch_related("addresses")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = listname

    headers = [
        'Bedrijfsnaam',
        'Ondernemingsnummer',
        'Adres',
        'Oprichting',
        'Website',
        'Juridische Vorm',
        'Eigen Vermogen',
        'Omzet',
        'Brutomarge',
        'EBITDA',
        'Winst/Verlies',
        'Netto Schuldpositie',
        'Capex Noden',
        'Bezoldigingen',
        'FTE',
        'Vastgoed'
    ]
    ws.append(headers)

    for company in companies:
        address = company.addresses.first()
        address_str = (
            f"{address.street} {address.house_number}, {address.postal_code} {address.city}"
            if address else ''
        )

        keyfigures = company.keyfigures or {}

        row = [
            company.name,
            company.number,
            address_str,
            company.start_date,
            company.website,
            company.legalform,
            keyfigures.get('equity', ''),
            keyfigures.get('turnover', ''),
            keyfigures.get('margin', ''),
            keyfigures.get('ebitda', ''),
            keyfigures.get('profit', ''),
            keyfigures.get('net_debt', ''),
            keyfigures.get('capex', ''),
            keyfigures.get('remuneration', ''),
            keyfigures.get('fte', ''),
            keyfigures.get('real_estate', '')
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
